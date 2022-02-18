# 輸入資料進Excel
import openpyxl
import datetime
from openpyxl.styles import Font, Alignment, PatternFill

class Input_to_Excel():
    def __init__(self, peopleNum):
        # 人數
        self.peopleNum = peopleNum

        # 例項化
        self.workbook = openpyxl.Workbook()

        # 啟用 worksheet
        self.sheet = self.workbook.active

        # 改工作表名稱
        self.sheet.title = "Present_Result"

        # 顏色
        self.color = ['00FFFFFF','00008000', "00C0C0C0"] # 白綠灰

        # 字體
        self.title1_font = Font(name=u"微軟正黑體", size=20, color=self.color[0], italic=False, bold=True)
        self.title2_font = Font(name=u"微軟正黑體", size=18, color=self.color[0], italic=False, bold=True)
        self.list_font = Font(name=u"微軟正黑體", size=12, italic=False, bold=False)

        # 填入背景顏色
        self.title_filled = PatternFill('solid', fgColor=self.color[1])
        self.list_gray_filled = PatternFill('solid', fgColor=self.color[2])

    # 寫入資料
    def inputData(self, nameList, presentList):
        # Title
        # A : 日期
        self.sheet['A1'] = "Date"

        # B : ID
        self.sheet['B1'] = "ID"

        # C : 名字
        self.sheet['C1'] = "Name"

        # D : 出席結果
        self.sheet['D1'] = "Present result"

        # Data
        for x in range(2, self.peopleNum + 2):
            # A : 日期
            self.sheet['A{}'.format(str(x))] = datetime.datetime.now().strftime("%Y-%m-%d")

            # B : ID
            self.sheet['B{}'.format(str(x))] = x - 1

            # C : 名字
            self.sheet['C{}'.format(str(x))] = nameList[x - 2]

            # D : 出席結果
            if presentList[x - 2] == 1:
                self.sheet['D{}'.format(str(x))] = "OK"
            else:
                self.sheet['D{}'.format(str(x))] = "Absence"

    # 裝飾表格
    def decorate_table(self):

        # 列寬
        self.sheet.column_dimensions['A'].width = 28
        self.sheet.column_dimensions['B'].width = 18
        self.sheet.column_dimensions['C'].width = 28
        self.sheet.column_dimensions['D'].width = 28

        # 行高
        for x in range(1, self.peopleNum + 20):
            if x == 1:
                self.sheet.row_dimensions[x].height = 30
            else:
                self.sheet.row_dimensions[x].height = 20

        # Title : 字體、置中和填充背景
        self.sheet['A1'].font = self.title1_font
        self.sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.sheet['A1'].fill = self.title_filled

        self.sheet['B1'].font = self.title1_font
        self.sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
        self.sheet['B1'].fill = self.title_filled

        self.sheet['C1'].font = self.title1_font
        self.sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
        self.sheet['C1'].fill = self.title_filled

        self.sheet['D1'].font = self.title2_font
        self.sheet['D1'].alignment = Alignment(horizontal='center', vertical='center')
        self.sheet['D1'].fill = self.title_filled

        # Data : 字體、置中和填充背景
        for x in range(2, self.peopleNum + 2):
            self.sheet['A{}'.format(str(x))].font = self.list_font
            self.sheet['A{}'.format(str(x))].alignment = Alignment(horizontal='center', vertical='center')
            if x % 2 == 0:
                self.sheet['A{}'.format(str(x))].fill = self.list_gray_filled


            self.sheet['B{}'.format(str(x))].font = self.list_font
            self.sheet['B{}'.format(str(x))].alignment = Alignment(horizontal='center', vertical='center')
            if x % 2 == 0:
                self.sheet['B{}'.format(str(x))].fill = self.list_gray_filled


            self.sheet['C{}'.format(str(x))].font = self.list_font
            self.sheet['C{}'.format(str(x))].alignment = Alignment(horizontal='center', vertical='center')
            if x % 2 == 0:
                self.sheet['C{}'.format(str(x))].fill = self.list_gray_filled


            self.sheet['D{}'.format(str(x))].font = self.list_font
            self.sheet['D{}'.format(str(x))].alignment = Alignment(horizontal='center', vertical='center')
            if x % 2 == 0:
                self.sheet['D{}'.format(str(x))].fill = self.list_gray_filled
                
    # 儲存Excel
    def save_Excel(self, excel_name):
        self.workbook.save(excel_name + ".xlsx")